import os
import openpyxl

# Dictionary mapping code -> descriptive language name.
# Used for reversing from descriptive name back to the code.
LANGUAGE_CODES = {
    "(chi-CN)": "Chinese (China)",
    "(chi-TW)": "Chinese (Taiwan)",
    "(chi-HK)": "Chinese (Hong-Kong)",
    "(cze)":    "Czech",
    "(dan)":    "Danish",
    "(dut-NL)": "Dutch (Netherlands)",
    "(dut-BE)": "Dutch (Belgium)",
    "(eng-GB)": "English",
    "(eng-US)": "English(United-States)",
    "(est)":    "Estonian",
    "(fin)":    "Finnish",
    "(fre-FR)": "French",
    "(fre-CA)": "French(Canada)",
    "(fre-CH)": "French(Switzerland)",
    "(fre-BE)": "French(Belgium)",
    "(ger-DE)": "German",
    "(ger-AT)": "German(Austria)",
    "(ger-CH)": "German(Switzerland)",
    "(hun)":    "Hungarian",
    "(ita-IT)": "Italian",
    "(jpn)":    "Japanese",
    "(kor)":    "Korean",
    "(lav)":    "Latvian",
    "(lit)":    "Lithuanian",
    "(nob)":    "Norwegian",
    "(pol)":    "Polish",
    "(por-BR)": "Portuguese (Brazil)",
    "(por-PT)": "Portuguese (Portugal)",
    "(rum)":    "Romanian",
    "(rus)":    "Russian",
    "(slo)":    "Slovak",
    "(slv)":    "Slovenian",
    "(spa-ES)": "Spanish",
    "(swe-SE)": "Swedish",
    "(tur)":    "Turkish",
    # Add additional mappings as needed
}

# Build a reverse lookup: "English" -> "(eng-GB)", etc.
REVERSE_LANGUAGE = {full: code for code, full in LANGUAGE_CODES.items()}

def create_en_files_from_excel(excel_path, output_folder):
    """
    Reads an Excel file (with the first column as 'Key' and subsequent columns 
    as languages) and creates individual .en files from each language column.

    Steps:
      1) Open the Excel file and read the header row, expecting at least:
          Key | <Language1> | <Language2> | ...
      2) For each row, store the Key -> string mapping per language.
      3) For each language column, output a single .en file with lines like:
           <KEY> = "<TEXT>"
         If multiple lines exist, wrap them as:
           <KEY> = [
           line1
           line2
           ]
      4) If the column header matches a recognized language name, 
         revert to the original code (e.g., '(eng-GB)').
         Otherwise, keep the header as-is in parentheses.
    """

    # Load the workbook and select the active sheet
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active

    # The first row is expected to be column headers
    headers = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    if len(headers) < 2:
        print("[ERROR] Excel must have at least two columns: 'Key' and one language column.")
        return

    # The first column is 'Key', others are language columns
    language_headers = headers[1:]
    translations_per_language = {lang: {} for lang in language_headers}

    # Iterate from row 2 onward, reading each row's key and language values
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            # Skip empty rows or rows with no key
            continue

        key = str(row[0]).strip()
        for idx, lang in enumerate(language_headers):
            cell_value = row[idx + 1]
            text = str(cell_value).rstrip() if cell_value else ""
            translations_per_language[lang][key] = text

    # Create output folder if needed
    os.makedirs(output_folder, exist_ok=True)

    # Base name for the .en files is the Excel filename minus extension
    base_name = os.path.splitext(os.path.basename(excel_path))[0]

    # For each language column, write a .en file
    for lang, translations_dict in translations_per_language.items():
        # If we recognize this language name, use the original code in parentheses
        if lang in REVERSE_LANGUAGE:
            lang_code = REVERSE_LANGUAGE[lang]  # e.g. "(eng-GB)"
            file_name = f"{base_name}{lang_code}.en"
        else:
            # If header is already in parentheses, use it directly; else wrap with parentheses
            if isinstance(lang, str) and lang.startswith("(") and lang.endswith(")"):
                file_name = f"{base_name}{lang}.en"
            else:
                file_name = f"{base_name}({lang}).en"

        file_path = os.path.join(output_folder, file_name)

        # Write each <KEY> in the dictionary to the file
        with open(file_path, "w", encoding="utf-8") as f_out:
            for key, value in translations_dict.items():
                lines = value.splitlines()
                if len(lines) > 1:
                    # Multi-line text
                    f_out.write(f"{key} = [\n")
                    for line in lines:
                        f_out.write(f"{line}\n")
                    f_out.write("]\n\n")
                else:
                    # Single-line text
                    single_line = lines[0] if lines else ""
                    f_out.write(f"{key} = \"{single_line}\"\n")

        print(f"[OK] Created: {file_path}")

def main():
    """
    This script expects:
      - An Excel file in the 'i' folder, e.g. "MyTranslations.xlsx"
      - A folder 'o' where generated .en files will be placed.
    
    By default, it looks for:
      i/CorrigoEVentilationTexts_5.3_complete.xlsx
    If found, it processes each language column, generating one .en file per column
    in the 'o' folder.
    """

    excel_file = os.path.join("i", "CorrigoEVentilationTexts_5.3_complete.xlsx")
    output_folder = "o"

    # Confirm the .xlsx file exists
    if not os.path.isfile(excel_file):
        print(f"[ERROR] Excel file '{excel_file}' not found.")
        return

    create_en_files_from_excel(excel_file, output_folder)

if __name__ == "__main__":
    main()
