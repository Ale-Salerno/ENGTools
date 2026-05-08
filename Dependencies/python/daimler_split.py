import os
import shutil
import json
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

def process_excel_files():
    # Use the current folder
    folder_path = os.getcwd()
    
    # Define target folders
    original_folder = os.path.join(folder_path, "original_metadata")
    split_folder = os.path.join(folder_path, "split_files")
    
    # Create folders if they don't exist
    os.makedirs(original_folder, exist_ok=True)
    os.makedirs(split_folder, exist_ok=True)
    
    overall_meta = {"files": []}
    
    # Process each .xlsx file in the folder
    for file_name in os.listdir(folder_path):
        if file_name.endswith(".xlsx"):
            file_path = os.path.join(folder_path, file_name)
            try:
                wb_source = load_workbook(file_path)
            except Exception as e:
                print(f"Error opening {file_name}: {e}")
                continue
            
            ws_source = wb_source.active  # Use the first sheet
            
            # --- Dynamic source column selection based on header "en" ---
            src_col_letter = None
            src_col_num = -1
            
            # Find the column with "en" in its header (case-insensitive)
            for col_idx in range(1, ws_source.max_column + 1):
                header_value = ws_source.cell(row=1, column=col_idx).value
                if header_value and str(header_value).strip().lower() == "en":
                    src_col_num = col_idx
                    src_col_letter = get_column_letter(src_col_num)
                    break
            
            if src_col_num == -1:
                print(f"Skipping {file_name}: No 'en' header found. Please ensure one column has 'en' as its header.")
                wb_source.close()
                continue

            # Hard-coded excluded columns: A (1) and B (2)
            exclude_cols = [1, 2]
            
            # Determine the used range (max_row and max_column)
            last_row = ws_source.max_row
            last_col = ws_source.max_column
            
            # Get the source locale from the header of the source column (row 1)
            source_locale = ws_source.cell(row=1, column=src_col_num).value
            
            file_meta = {
                "originalFile": file_name,
                "sourceColumn": src_col_letter,
                "excludedColumns": ["A", "B"], # These are hardcoded as per your original script
                "mappings": []
            }
            
            # Loop through columns in the sheet
            for col in range(1, last_col + 1):
                # Skip source column and excluded columns
                if col == src_col_num or col in exclude_cols:
                    continue
                
                # Get target locale from header (row 1)
                target_locale = ws_source.cell(row=1, column=col).value
                current_col_letter = get_column_letter(col)
                
                # Create a new workbook with one sheet
                wb_new = Workbook()
                ws_new = wb_new.active
                
                # Write header row in the new workbook
                ws_new.cell(row=1, column=1, value=source_locale)
                ws_new.cell(row=1, column=2, value=target_locale)
                
                # Copy data rows from row 2 to last_row
                for r in range(2, last_row + 1):
                    ws_new.cell(row=r, column=1, value=ws_source.cell(row=r, column=src_col_num).value)
                    ws_new.cell(row=r, column=2, value=ws_source.cell(row=r, column=col).value)
                
                # Construct the new file name:
                # originalfilename_sourcelocale_targetlocale.xlsx
                base_name, _ = os.path.splitext(file_name)
                # Sanitize locale names for filenames (replace problematic characters)
                sanitized_source_locale = "".join([c for c in source_locale if c.isalnum() or c in ('_', '-')]).strip() if source_locale else "unknown"
                sanitized_target_locale = "".join([c for c in target_locale if c.isalnum() or c in ('_', '-')]).strip() if target_locale else "unknown"
                
                new_file_name = f"{base_name}_{sanitized_source_locale}_{sanitized_target_locale}.xlsx"
                new_file_path = os.path.join(split_folder, new_file_name)
                
                wb_new.save(new_file_path)
                
                # Append mapping metadata for this target column
                mapping = {
                    "targetColumn": current_col_letter,
                    "sourceLocale": source_locale,
                    "targetLocale": target_locale,
                    "generatedFile": new_file_name
                }
                file_meta["mappings"].append(mapping)
            
            overall_meta["files"].append(file_meta)
            wb_source.close()
            
            # Move original file to the original_metadata folder
            shutil.move(file_path, os.path.join(original_folder, file_name))
    
    # Save metadata as JSON in the original_metadata folder
    meta_file_path = os.path.join(original_folder, "metadata_conversion.json")
    with open(meta_file_path, "w", encoding="utf-8") as f:
        json.dump(overall_meta, f, indent=4)
    
    print("Processing completed.")
    print(f"Metadata saved to {meta_file_path}")

if __name__ == "__main__":
    process_excel_files()