import pandas as pd
import os
import sys
import json
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

# --- Metadata generation ---
def generate_merge_metadata(source_file, source_col, target_col, bilingual_file, sheet_hidden_rows, sheet_visible_rows):
    """Record relationships for merging"""
    meta_file = os.path.join(script_dir, "excel_merge_metadata.json")
    if os.path.exists(meta_file):
        with open(meta_file, 'r', encoding='utf-8') as f:
            metadata = json.load(f)
    else:
        metadata = {"files": []}

    entry = {
        "source_file": source_file,
        "original_columns": {
            "source": index_to_column(source_col),
            "target": index_to_column(target_col)
        },
        "hidden_rows": sheet_hidden_rows,
        "visible_rows": sheet_visible_rows,
        "bilingual_file": bilingual_file,
        "created_at": datetime.now().isoformat()
    }
    if entry not in metadata["files"]:
        metadata["files"].append(entry)
        with open(meta_file, 'w', encoding='utf-8') as f:
            json.dump(metadata, f, indent=2)

# --- Column utilities ---
def column_to_index(col_letter):
    """Convert Excel column letter (e.g. "A", "AA") to 0-based index."""
    return column_index_from_string(col_letter) - 1

def index_to_column(index):
    """Convert 0-based column index to Excel column letter."""
    return get_column_letter(index + 1)

# --- Hidden rows detection ---
def get_hidden_rows(ws):
    """Get 0-based indices of hidden rows in a worksheet"""
    return [r - 1 for r, dim in ws.row_dimensions.items() if dim.hidden]

# --- Header finder ---
def find_header_row(df, source_idx, target_idx, max_rows=20):
    if df.empty:
        return None
    for row_idx in range(min(max_rows, len(df))):
        try:
            if pd.notna(df.iat[row_idx, source_idx]) and pd.notna(df.iat[row_idx, target_idx]):
                return row_idx
        except IndexError:
            return None
    return None

# --- Main processing ---
def main():
    source_col = input("Enter source column letter (e.g. A, B, C): ").strip()
    if not source_col:
        print("Error: No source column provided. Exiting.")
        sys.exit(1)
    exclude_input = input("Enter columns to exclude (comma-separated, optional): ").strip()
    exclude_cols = [c.strip() for c in exclude_input.split(',')] if exclude_input else []
    try:
        source_idx = column_to_index(source_col)
    except Exception as e:
        print(f"Invalid source column '{source_col}': {e}")
        sys.exit(1)
    try:
        exclude_indices = [column_to_index(c) for c in exclude_cols]
    except Exception as e:
        print(f"Invalid exclude columns '{exclude_input}': {e}")
        sys.exit(1)

    global script_dir
    script_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_files = [f for f in os.listdir(script_dir) if f.endswith('.xlsx') and not f.startswith('~$')]

    for filename in xlsx_files:
        file_path = os.path.join(script_dir, filename)
        base_name = os.path.splitext(filename)[0]
        backup_path = os.path.join(script_dir, f"{base_name}_backup.xlsx")
        if not os.path.exists(backup_path):
            shutil.copyfile(file_path, backup_path)
            print(f"Backup created: {os.path.basename(backup_path)}")

        # Detect hidden rows per sheet
        wb = load_workbook(file_path)
        sheet_names = wb.sheetnames
        sheet_hidden_rows = {}
        sheet_visible_rows = {}
        for name in sheet_names:
            ws = wb[name]
            hidden = get_hidden_rows(ws)
            sheet_hidden_rows[name] = hidden
            sheet_visible_rows[name] = [i for i in range(ws.max_row) if i not in hidden]
        wb.close()

        # Determine valid, visible columns (exclude hidden)
        wb_tmp = load_workbook(file_path)  # need full workbook to access column_dimensions
        max_cols = set()
        hidden_cols = set()
        for ws in wb_tmp.worksheets:
            max_cols.update(range(ws.max_column))
            for letter, dim in ws.column_dimensions.items():
                if dim.hidden:
                    hidden_cols.add(column_to_index(letter))
        wb_tmp.close()
        valid_columns = max_cols - hidden_cols

        if source_idx not in valid_columns:
            print(f"Skipping {filename}: Invalid or hidden source column {source_col}")
            continue

        targets = valid_columns - {source_idx} - set(exclude_indices)
        for target_idx in sorted(targets):
            has_data = False
            src_header = tgt_header = None
            for name in sheet_names:
                df = pd.read_excel(file_path, sheet_name=name, header=None, skiprows=sheet_hidden_rows[name])
                header_row = find_header_row(df, source_idx, target_idx)
                start = header_row + 1 if header_row is not None else 0
                try:
                    col = df.iloc[start:, target_idx]
                except IndexError:
                    continue
                if col.notna().any():
                    has_data = True
                    if header_row is not None:
                        src_header = str(df.iat[header_row, source_idx])
                        tgt_header = str(df.iat[header_row, target_idx])
                    break
            if not has_data:
                continue
            src_header = src_header or index_to_column(source_idx)
            tgt_header = tgt_header or index_to_column(target_idx)
            out_name = f"{base_name}_{src_header}_{tgt_header}.xlsx"
            out_path = os.path.join(script_dir, out_name)
            generate_merge_metadata(filename, source_idx, target_idx, out_name, sheet_hidden_rows, sheet_visible_rows)
            with pd.ExcelWriter(out_path) as writer:
                for name in sheet_names:
                    df = pd.read_excel(file_path, sheet_name=name, header=None, skiprows=sheet_hidden_rows[name])
                    header_row = find_header_row(df, source_idx, target_idx)
                    if header_row is not None:
                        headers = [df.iat[header_row, source_idx], df.iat[header_row, target_idx]]
                        data = df.iloc[header_row + 1:, [source_idx, target_idx]]
                    else:
                        headers = [index_to_column(source_idx), index_to_column(target_idx)]
                        data = df.iloc[:, [source_idx, target_idx]]
                    pd.DataFrame(data.values, columns=headers).to_excel(writer, sheet_name=name, index=False)
    print("Processing complete! Hidden rows ignored.")

if __name__ == "__main__":
    main()