import os
import re
import datetime
import getpass

import pandas as pd
import openpyxl
from langdetect import detect, DetectorFactory
import xml.etree.ElementTree as ET
from xml.dom import minidom

# Set a seed to make langdetect deterministic across runs
DetectorFactory.seed = 0

LANG_CODE_MAPPING = {
    "en": "en-US", "fr": "fr-FR", "es": "es-ES", "de": "de-DE", "it": "it-IT",
    "pt": "pt-PT", "pt-br": "pt-BR", "nl": "nl-NL", "af": "af-ZA", "ru": "ru-RU",
    "zh-cn": "zh-CN", "zh-tw": "zh-TW", "ja": "ja-JP", "ko": "ko-KR", "ar": "ar-SA",
    "tr": "tr-TR", "sv": "sv-SE", "fi": "fi-FI", "da": "da-DK", "no": "no-NO", "pl": "pl-PL",
    "cs": "cs-CZ", "hu": "hu-HU", "el": "el-GR", "bg": "bg-BG", "mk": "mk-MK", "et": "et-EE",
    "sl": "sl-SI", "lt": "lt-LT", "lv": "lv-LV", "ro": "ro-RO", "sk": "sk-SK", "hr": "hr-HR",
    "uk": "uk-UA", "sq": "sq-AL", "sr": "sr-RS", "hy": "hy-AM", "ka": "ka-GE",
    "bn": "bn-BD", "mr": "mr-IN", "gu": "gu-IN", "ta": "ta-IN", "te": "te-IN", "ml": "ml-IN",
    "kn": "kn-IN", "pa": "pa-IN", "or": "or-IN", "as": "as-IN", "ne": "ne-NP", "si": "si-LK",
    "km": "km-KH", "my": "my-MM", "lo": "lo-LA", "th": "th-TH", "vi": "vi-VN", "id": "id-ID",
    "ms": "ms-MY", "fil": "fil-PH", "cy": "cy-GB", "is": "is-IS",
    "eu": "eu-ES", "ca": "ca-ES", "gl": "gl-ES", "fa": "fa-IR", "ur": "ur-PK", "sw": "sw-KE",
    "he": "he-IL", "iw": "iw-IL", "zh": "zh-CN", "pt-PT": "pt-PT", "pt-BR": "pt-BR",
    "bg-BG": "bg-BG"
}

def main():
    """
    This script processes one or more Excel files (all .xlsx / .xls in the current directory)
    to create TMX files (Translation Memory eXchange). For each workbook:

      1) It iterates through all sheets and attempts to detect or prompt for the
         source language column.
      2) For each other column in the sheet, attempts to detect or prompt for a target language.
      3) Creates a pairwise TMX for each source-target column pair.
      4) If more than one target language column is present, also creates a single,
         multilingual TMX combining them all.

    Steps Involved:
      - Detecting languages by either analyzing the column header or sampling text within it.
      - If detection is inconclusive or duplicates appear, the user is prompted to confirm or
        specify a language code.
      - Extracts text from each cell, optionally splitting lines by soft returns (Alt+Enter).
      - Writes each row's text to a <tu> in a TMX file, along with metadata (filename, sheet name,
        and cell positions).

    Usage:
      - Place this script in the same directory as your Excel files.
      - Run the script from that directory.
      - Follow on-screen prompts to select source columns and confirm or override language codes.
      - Outputs .tmx files per language combination and possibly a multilingual file if multiple
        targets exist.
    """

    excel_files = [file for file in os.listdir() if file.lower().endswith((".xlsx", ".xls"))]

    if not excel_files:
        print("[ERROR] No Excel files found in the current directory.")
        return

    creation_id = f"{getpass.getuser()}@languagewire.com"

    # Process each Excel file found.
    for master_file in excel_files:
        print(f"\nReading Excel file: {master_file}")

        try:
            # Load all sheets into a dictionary of DataFrames: {sheet_name: DataFrame}
            dfs = pd.read_excel(master_file, sheet_name=None)
        except Exception as e:
            print(f"[ERROR] Could not read '{master_file}': {e}")
            continue

        for sheet_name, sheet_df in dfs.items():
            if sheet_df.empty or len(sheet_df.columns) < 2:
                print(f"[SKIP] Sheet '{sheet_name}' is empty or has fewer than 2 columns.")
                continue

            # Prompt the user for the source column letter (defaults to "A").
            user_input = input(
                f"Enter the source column letter for sheet '{sheet_name}' (default A): "
            ).strip()
            source_col_letter = user_input.upper() if user_input else "A"

            # Convert the column letter to an index, ensuring it is in range.
            source_index = column_letter_to_index(source_col_letter)
            if source_index >= len(sheet_df.columns):
                print(f"[WARNING] Column '{source_col_letter}' is out of range for sheet '{sheet_name}'. Using column A.")
                source_index = 0
                source_col_letter = "A"

            # Attempt to detect the language for the chosen source column.
            source_header = str(sheet_df.columns[source_index]).strip()
            # Find the first non-empty cell in the source column as a sample.
            source_sample = next((txt for txt in sheet_df.iloc[:, source_index] if pd.notna(txt)), "")
            source_lang = detect_language(source_header, source_sample) or LANG_CODE_MAPPING.get("en")

            print(f"Detected source language for sheet '{sheet_name}': {source_lang}")

            # Next, detect languages for the other columns and handle duplicates or confusion.
            target_langs = {}
            detected_languages = {}

            for col in sheet_df.columns:
                if col == sheet_df.columns[source_index]:
                    continue
                target_header = str(col).strip()
                # Sample up to five non-empty cells to detect language.
                target_sample_texts = list(sheet_df[col].dropna().astype(str))[:5]
                first_sample = target_sample_texts[0] if target_sample_texts else ""
                detected_languages[col] = detect_language(target_header, first_sample)

            # Check if multiple columns share the same detected language.
            duplicated_languages_columns = {}
            for col, lang in detected_languages.items():
                if lang in LANG_CODE_MAPPING.values():
                    duplicated_languages_columns.setdefault(lang, []).append(col)

            # If a language is detected in multiple columns, ask user to reassign them distinctly.
            for lang, cols in duplicated_languages_columns.items():
                if len(cols) > 1:
                    print(f"\nLanguage '{lang}' was detected in multiple columns: {', '.join(cols)}")
                    for col in cols:
                        target_header = str(col).strip()
                        target_sample_texts = list(sheet_df[col].dropna().astype(str))[:5]
                        chosen_lang = prompt_for_language(target_header, target_sample_texts, target_langs.values())
                        target_langs[col] = chosen_lang
                else:
                    # Only one column found with this detected language -> no conflict
                    target_langs[cols[0]] = lang

            # If any columns are still missing a language assignment, prompt the user
            for col in sheet_df.columns:
                if col == sheet_df.columns[source_index]:
                    continue
                if col not in target_langs:
                    target_header = str(col).strip()
                    target_sample_texts = list(sheet_df[col].dropna().astype(str))[:5]
                    chosen_lang = prompt_for_language(target_header, target_sample_texts, target_langs.values())
                    target_langs[col] = chosen_lang

                print(f"Assigned target language '{target_langs[col]}' to column '{col}'.")

            # Build a dictionary with the source column language + all target columns
            language_mapping = {sheet_df.columns[source_index]: source_lang}

            # Create pairwise TMX files for each target column
            for col in sheet_df.columns:
                if col == sheet_df.columns[source_index]:
                    continue

                target_lang = target_langs[col]
                language_mapping[col] = target_lang

                # Skip if the target language is the same as the source language.
                if target_lang == source_lang:
                    print(f"[SKIP] Column '{col}' has the same language as the source ({target_lang}). No TMX created.")
                    continue

                source_texts = sheet_df.iloc[:, source_index]
                target_texts = sheet_df[col]

                # 1-based column index for Excel
                col_letter_index = sheet_df.columns.get_loc(col)
                column_letter_target = get_column_letter(col_letter_index + 1)

                output_file_name = (
                    f"{os.path.splitext(master_file)[0]}_"
                    f"{sheet_name}_{source_lang}_{target_lang}.tmx"
                )

                create_tmx(
                    source_lang=source_lang,
                    target_lang=target_lang,
                    source_texts=source_texts,
                    target_texts=target_texts,
                    output_file=output_file_name,
                    creation_id=creation_id,
                    filename=master_file,
                    sheet_name=sheet_name,
                    column_letter_source=source_col_letter,
                    column_letter_target=column_letter_target
                )
                print(f"[OK] Created TMX file: {output_file_name}")

            # Optionally create a single multilingual TMX if more than one target language.
            source_col = sheet_df.columns[source_index]
            new_order = [source_col] + [c for c in sheet_df.columns if c != source_col]
            sheet_df_reordered = sheet_df[new_order]

            if len(target_langs) > 1:
                multilingual_tmx_file = f"{os.path.splitext(master_file)[0]}_{sheet_name}_multilingual.tmx"
                create_multilingual_tmx(
                    sheet_df_reordered,
                    multilingual_tmx_file,
                    creation_id,
                    master_file,
                    language_mapping,
                    sheet_name
                )
                print(f"[OK] Created multilingual TMX file: {multilingual_tmx_file}")
            else:
                print(f"[SKIP] Only one or zero target languages for '{sheet_name}'. No multilingual TMX generated.")


# --------------------- Below Are Helper Functions --------------------- #

def is_valid_lang_code(code):
    """Return True if 'code' is present in LANG_CODE_MAPPING keys (case-sensitive check)."""
    return code in LANG_CODE_MAPPING

def detect_language(header, sample_text):
    """
    Attempt to detect a language code by:
      1) Checking if 'header' (lowercased) can be mapped directly.
      2) If not, splitting on '-' to see if the prefix is recognized (e.g. 'de' from 'de-DE').
      3) If still not found, use langdetect on 'sample_text' to guess a language code.
    Returns the mapped code (e.g., 'en-US') if found, otherwise None.
    """
    header_lower = header.lower().strip()
    mapping_lower = {k.lower(): v for k, v in LANG_CODE_MAPPING.items()}

    # Direct lookup
    if header_lower in mapping_lower:
        header_lang = mapping_lower[header_lower]
        try:
            detected = detect(sample_text)
            detected_lang = mapping_lower.get(detected.lower().strip())
            if detected_lang and detected_lang != header_lang:
                print(f"[WARNING] Header language '{header_lang}' != detected language '{detected_lang}'. Using header language.")
        except Exception:
            print("[WARN] Could not detect language from sample text; using header language.")
        return header_lang

    # Attempt prefix-based detection, e.g. 'de' from 'de-DE'
    if '-' in header_lower:
        prefix = header_lower.split('-', 1)[0]
        if prefix in mapping_lower:
            return mapping_lower[prefix]

    # Fall back to langdetect if header-based detection fails
    try:
        detected = detect(sample_text)
        return mapping_lower.get(detected.lower().strip(), None)
    except Exception:
        return None

def prompt_for_language(header, sample_texts, existing_langs=None):
    """
    Prompt user for a valid language code. Show up to 5 sample cells from the column.
    Reject duplicates if a code is already in 'existing_langs'.
    """
    print(f"\nNo clear language code for column: '{header}'")
    print("Sample text from this column:")
    for i, text in enumerate(sample_texts[:5]):
        print(f"  {i+1}. {text}")

    if existing_langs:
        print(f"\nLanguages already used: {', '.join(existing_langs)}")

    while True:
        lang_code = input("Enter a valid language code (e.g. en-US, fr-FR, pt-BR): ").strip()
        if is_valid_lang_code(lang_code) and (lang_code not in (existing_langs or [])):
            return lang_code
        elif existing_langs and lang_code in existing_langs:
            print("[WARN] This code is already used. Please enter a different code.")
        else:
            print("[WARN] Invalid code. Try again.")

def clean_text(text):
    """Convert non-string or NaN to empty, strip whitespace, remove extra newlines."""
    if pd.isna(text):
        return ""
    text = str(text).strip()
    text = re.sub(r'\s+', ' ', text)
    text = text.replace('\r', '')
    return text

def get_column_letter(col_index):
    """
    Convert 1-based col_index to an Excel column letter.
    e.g. 1 -> 'A', 2 -> 'B', 27 -> 'AA'.
    """
    letter = ''
    while col_index > 0:
        col_index, remainder = divmod(col_index - 1, 26)
        letter = chr(65 + remainder) + letter
    return letter

def column_letter_to_index(letter):
    """Convert an Excel column letter to zero-based index (e.g. 'A'->0, 'B'->1, 'AA'->26)."""
    letter = letter.upper()
    index = 0
    for char in letter:
        if 'A' <= char <= 'Z':
            index = index * 26 + (ord(char) - ord('A') + 1)
    return index - 1

def get_cell_position(row_index, col_index):
    """
    Return an Excel-style cell reference (e.g. 'B2') given a zero-based row and col index.
    Adds 2 to row since row 0 = header row, row 1 = first data row in the sheet, etc.
    """
    col_letter = get_column_letter(col_index)
    return f"{col_letter}{row_index + 2}"

def split_by_soft_return(text):
    """Split the string by newline characters (representing Alt+Enter in Excel)."""
    if pd.notna(text):
        return str(text).split('\n')
    return [""]

def create_tmx(source_lang, target_lang, source_texts, target_texts, output_file,
               creation_id, filename, sheet_name, column_letter_source, column_letter_target):
    """
    Create a bilingual TMX file using source_texts and target_texts from a single Excel sheet.
    Each row in the Excel column is written to a <tu> in the TMX. Also includes metadata
    like filename, sheet, and cell positions.
    """
    # Attempt loading the workbook/sheet for additional context (optional).
    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb[sheet_name]
    except Exception:
        sheet = None

    # Build the basic TMX structure.
    root = ET.Element("tmx", version="1.4")
    header_attribs = {
        "creationtool": "Excel2TMX",
        "creationtoolversion": "1.0",
        "datatype": "PlainText",
        "segtype": "sentence",
        "adminlang": "en-US",
        "srclang": source_lang,
        "o-tmf": "Excel2TMX"
    }
    ET.SubElement(root, "header", header_attribs)
    body = ET.SubElement(root, "body")

    seen_segments = {}

    for idx, (src_cell, tgt_cell) in enumerate(zip(source_texts, target_texts)):
        src_lines = split_by_soft_return(src_cell)
        tgt_lines = split_by_soft_return(tgt_cell)
        max_lines = max(len(src_lines), len(tgt_lines))

        for line_idx in range(max_lines):
            src_cleaned = clean_text(src_lines[line_idx] if line_idx < len(src_lines) else "")
            tgt_cleaned = clean_text(tgt_lines[line_idx] if line_idx < len(tgt_lines) else "")
            if not src_cleaned or not tgt_cleaned:
                continue

            segment_key = src_cleaned

            tu = ET.SubElement(
                body, "tu",
                creationid=creation_id,
                creationdate=datetime.datetime.now().isoformat()
            )
            prop_filename = ET.SubElement(tu, "prop", {"type": "filename"})
            prop_filename.text = filename
            prop_sheetname = ET.SubElement(tu, "prop", {"type": "sheet"})
            prop_sheetname.text = sheet_name

            # Build source/target cell references if the Excel sheet was successfully loaded.
            if sheet is not None:
                source_position = f"{column_letter_source}{idx + 2}"
                target_position = f"{column_letter_target}{idx + 2}"
            else:
                source_position = target_position = ""

            prop_position_src = ET.SubElement(tu, "prop", {"type": "source_position"})
            prop_position_src.text = source_position
            prop_position_tgt = ET.SubElement(tu, "prop", {"type": "target_position"})
            prop_position_tgt.text = target_position

            # <tuv> for source language
            tuv_src = ET.SubElement(tu, "tuv", {"xml:lang": source_lang})
            seg_src = ET.SubElement(tuv_src, "seg")
            seg_src.text = src_cleaned

            # <tuv> for target language
            tuv_tgt = ET.SubElement(tu, "tuv", {"xml:lang": target_lang})
            seg_tgt = ET.SubElement(tuv_tgt, "seg")
            seg_tgt.text = tgt_cleaned

            # If we have seen the same source segment before, replace it with this new entry.
            if segment_key not in seen_segments:
                seen_segments[segment_key] = tu
            else:
                body.remove(seen_segments[segment_key])
                seen_segments[segment_key] = tu

    # Pretty-print the XML and save to file.
    rough_string = ET.tostring(root, encoding='utf-8')
    reparsed = minidom.parseString(rough_string)
    pretty_xml = reparsed.toprettyxml(indent="  ")

    try:
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(pretty_xml)
    except (OSError, IOError) as e:
        print(f"[ERROR] Could not write '{output_file}': {e}")

def create_multilingual_tmx(data, output_file, creation_id, filename, language_mapping, sheet_name):
    """
    Build a single TMX file containing more than two languages from the provided DataFrame columns.
    'data' is expected to have its first column as the source, with additional columns as targets.

    The 'language_mapping' dict maps each column to a recognized language code (e.g., 'en-US').

    Each row is split by soft returns into lines. Each line becomes one <tu>.
    For each language column, a <tuv xml:lang="..."> is added with the cleaned text.
    """
    # Root TMX element
    root = ET.Element("tmx", version="1.4")
    header_attribs = {
        "creationtool": "Excel2TMX",
        "creationtoolversion": "1.0",
        "datatype": "PlainText",
        "segtype": "sentence",
        "adminlang": "en-US",
        "srclang": language_mapping[data.columns[0]],
        "o-tmf": "Excel2TMX"
    }
    ET.SubElement(root, "header", header_attribs)
    body = ET.SubElement(root, "body")

    seen_segments = {}

    for row_index, row in data.iterrows():
        # The first column is the "source" in language_mapping, but each column is included.
        src_col = data.columns[0]
        src_lines = split_by_soft_return(row[src_col])
        max_lines = len(src_lines)

        for line_idx, src_line in enumerate(src_lines):
            src_cleaned = clean_text(src_line)
            if not src_cleaned:
                continue

            segment_key = src_cleaned  # used to detect duplicates

            tu = ET.SubElement(
                body, "tu",
                creationid=creation_id,
                creationdate=datetime.datetime.now().isoformat()
            )
            prop_filename = ET.SubElement(tu, "prop", {"type": "filename"})
            prop_filename.text = filename
            prop_sheetname = ET.SubElement(tu, "prop", {"type": "sheet"})
            prop_sheetname.text = sheet_name

            segment_texts = {}

            # For each column in the DataFrame, add a <tuv> if we have text.
            for col_index, col in enumerate(data.columns, start=1):
                lang_code = language_mapping[col]  # e.g. 'en-US'
                col_lines = split_by_soft_return(row[col])
                cleaned_col_line = clean_text(col_lines[line_idx] if line_idx < len(col_lines) else "")
                if not cleaned_col_line:
                    continue

                # <tuv> for each language
                tuv = ET.SubElement(tu, "tuv", {"xml:lang": lang_code})
                seg = ET.SubElement(tuv, "seg")
                seg.text = cleaned_col_line

                cell_position = get_cell_position(row_index, col_index)
                prop_position = ET.SubElement(tuv, "prop", {"type": "position"})
                prop_position.text = cell_position

                segment_texts[lang_code] = cleaned_col_line

            if segment_key not in seen_segments:
                seen_segments[segment_key] = {
                    "segment_texts": segment_texts,
                    "tu_element": tu
                }
            else:
                # If the same segment key is encountered, replace with the new <tu>.
                body.remove(seen_segments[segment_key]["tu_element"])
                seen_segments[segment_key] = {
                    "segment_texts": segment_texts,
                    "tu_element": tu
                }

    # Pretty-print final TMX content.
    rough_string = ET.tostring(root, encoding='utf-8')
    reparsed = minidom.parseString(rough_string)
    pretty_xml = reparsed.toprettyxml(indent="  ")
    try:
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(pretty_xml)
    except (OSError, IOError) as e:
        print(f"[ERROR] Could not write '{output_file}': {e}")


# ----------------------------- Script Entry Point ----------------------------- #
if __name__ == "__main__":
    main()
