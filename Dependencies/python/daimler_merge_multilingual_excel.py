import os
import json
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

def revert_conversion():
    # Use the current working directory as the folder path.
    folder_path = os.getcwd()
    
    # Path to the metadata file (created during the splitting process).
    metadata_path = os.path.join(folder_path, "metadata_conversion.json")
    
    # Check if the metadata file exists.
    if not os.path.exists(metadata_path):
        print(f"Metadata file not found: {metadata_path}")
        return

    # Load the metadata JSON.
    with open(metadata_path, "r", encoding="utf-8") as f:
        metadata = json.load(f)
    
    # Iterate over each file entry in the metadata.
    for file_entry in metadata.get("files", []):
        orig_filename = file_entry.get("originalFile")
        orig_filepath = os.path.join(folder_path, orig_filename)
        
        if not os.path.exists(orig_filepath):
            print(f"Original file not found: {orig_filepath}")
            continue
        
        print(f"Processing original file: {orig_filename}")
        # Open the original multilingual workbook.
        wb_orig = load_workbook(orig_filepath)
        ws_orig = wb_orig.active  # Assuming the first (active) sheet
        
        # For each mapping (each target language split), update the original file.
        for mapping in file_entry.get("mappings", []):
            target_column_letter = mapping.get("targetColumn")
            generated_filename = mapping.get("generatedFile")
            generated_filepath = os.path.join(folder_path, generated_filename)
            
            if not os.path.exists(generated_filepath):
                print(f"Generated file not found: {generated_filepath}")
                continue
            
            print(f"  Updating target column {target_column_letter} from file: {generated_filename}")
            # Open the generated target workbook.
            wb_gen = load_workbook(generated_filepath)
            ws_gen = wb_gen.active
            
            # Determine the last row in the generated workbook.
            last_row = ws_gen.max_row
            
            # For each row (including header) copy the value from column B of the generated file.
            for row in range(1, last_row + 1):
                value = ws_gen.cell(row=row, column=2).value  # Column B
                # Convert the target column letter to its numerical index.
                col_index = column_index_from_string(target_column_letter)
                # Paste the value into the corresponding cell in the original workbook.
                ws_orig.cell(row=row, column=col_index, value=value)
            
            wb_gen.close()  # Close the generated file.
        
        # Save the updated workbook as a new file with the suffix "_compiled".
        base_name, ext = os.path.splitext(orig_filename)
        new_filename = f"{base_name}_compiled{ext}"
        new_filepath = os.path.join(folder_path, new_filename)
        wb_orig.save(new_filepath)
        wb_orig.close()
        print(f"compiled file saved as: {new_filename}")

if __name__ == "__main__":
    revert_conversion()