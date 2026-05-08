import os
import json
from openpyxl import load_workbook
from openpyxl.styles import Protection
from openpyxl.utils import column_index_from_string

def merge_translations():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    meta_file = os.path.join(script_dir, "excel_merge_metadata.json")
    
    if not os.path.exists(meta_file):
        print("Error: No merge metadata found. Run the splitter first.")
        return

    with open(meta_file, encoding='utf-8') as f:
        metadata = json.load(f)

    for entry in metadata["files"]:
        source_path    = os.path.join(script_dir, entry["source_file"])
        bilingual_path = os.path.join(script_dir, entry["bilingual_file"])

        if not os.path.exists(bilingual_path):
            print(f"Skipping {entry['source_file']}: Bilingual file missing")
            continue

        src_wb = load_workbook(source_path)
        bil_wb = load_workbook(bilingual_path)
        
        src_letter = entry["original_columns"]["source"]
        tgt_letter = entry["original_columns"]["target"]
        src_col = column_index_from_string(src_letter)
        tgt_col = column_index_from_string(tgt_letter)

        for sheet_name in src_wb.sheetnames:
            if sheet_name not in bil_wb.sheetnames:
                continue

            src_ws = src_wb[sheet_name]
            bil_ws = bil_wb[sheet_name]

            # Skip if target column is hidden in the source sheet
            col_dim = src_ws.column_dimensions.get(tgt_letter)
            if col_dim and col_dim.hidden:
                print(f"Skipping hidden column {tgt_letter} in '{sheet_name}'")
                continue

            visible_rows = [r + 1 for r in entry["visible_rows"][sheet_name]]

            for idx, bil_row in enumerate(range(1, bil_ws.max_row + 1)):
                if idx >= len(visible_rows):
                    break

                src_row  = visible_rows[idx]
                bil_cell = bil_ws.cell(row=bil_row, column=2)
                tgt_cell = src_ws.cell(row=src_row, column=tgt_col)

                tgt_cell.value         = bil_cell.value
                tgt_cell.font          = bil_cell.font.copy()
                tgt_cell.border        = bil_cell.border.copy()
                tgt_cell.fill          = bil_cell.fill.copy()
                tgt_cell.number_format = bil_cell.number_format
                tgt_cell.protection    = Protection(locked=False)

        src_wb.save(source_path)
        print(f"Merged: {os.path.basename(source_path)}")

if __name__ == "__main__":
    merge_translations()
    print("Merge completed.")
